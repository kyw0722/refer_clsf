# coding:utf-8
"""
Excel 파일을 처리하는 다양한 함수 모음
"""
from openpyxl import load_workbook, Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re


def read_excel_to_list(excel_filename:str, sheet_index:int, field_idxs:list, min_row:int):
    """
    excel 파일의 내용을 읽어서 list로 만든다.
    :param excel_filename:
    :param sheet_index: sheet 의 index
    :param field_idxs: 컬럼의 index 순서 지정,  eq [0, 3, 1, None] 이라면 리스트에 0번째 컬럼, 3번째 컬럼, 1번째 컬럼 순으로 list만듬
    :param min_row:
    :return:
    """
    excel_obj = load_workbook(filename=excel_filename, read_only=True, data_only=True, keep_links=False)
    sheet_name = excel_obj.sheetnames[sheet_index]
    active_sheet = excel_obj[sheet_name]

    total_list = []
    for idx, row in enumerate(active_sheet.iter_rows(min_row=min_row)):
        oneitem_length = len(field_idxs)
        oneitem_list = [None for _ in range(oneitem_length)]

        for idx, field_idx in enumerate(field_idxs):
            oneitem_list[idx] = row[field_idx].value
        total_list.append(oneitem_list)

    return total_list


def write_excel_from_list_with_style(excel_filename, data_list, head_list):

    book = Workbook()
    sheet = book.active

    item_list = []

    for idx, v in enumerate(head_list):
        cell = WriteOnlyCell(sheet, v)
        cell.alignment = Alignment(vertical="center", wrap_text=True)  # text alignment. Alignment(horizontal=, vertical=)

        # if idx >= (len(head_list) - 6):
        #     cell.font = Font(color='0000FF', underline="single", bold=True)
        #     cell.fill = PatternFill(start_color='FFBB00', end_color='FFBB00', fill_type='solid')
        #     cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        item_list.append(cell)

    sheet.append(item_list)

    for idx, col in enumerate(sheet.columns):
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        sheet.column_dimensions[col_name].width = 20

    for data in data_list:
        sheet.append(data)

    book.save(excel_filename)


def write_excel_from_list_multi_sheet(excel_path, data_list, head_list, sheet_name_list):
    book = Workbook()
    sheet_cnt = len(data_list)
    for idx in range(sheet_cnt):
        sheet = book.create_sheet(title=sheet_name_list[idx], index=idx)
        sheet.append(head_list[idx])
        data = data_list[idx]
        for item in data:
            sheet.append(item)
    book.save(excel_path)
    book.close()
    print('{} 생성완료'.format(excel_path))

